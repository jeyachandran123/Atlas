"""
DIP Phase 5 — Intelligent Content Generation verification.

Covers: knowledge-grounded artifact generation across all 7 formats,
lifecycle + events + metrics, checksums, registry/manifest fields, signed
download (or proxy fallback), file-format validity of the downloaded bytes,
unknown-format rejection, and ownership/404 semantics.
"""
import hashlib
import io
import json
import sys
import time
import uuid

import httpx

BASE = "http://127.0.0.1:8000/api/v1"
EMAIL = f"p5-{uuid.uuid4().hex[:8]}@example.com"
PASSWORD = "p5-password-123"

PASS = 0
FAIL = 0


def check(name: str, cond: bool, detail: str = "") -> None:
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  PASS  {name}")
    else:
        FAIL += 1
        print(f"  FAIL  {name}  {detail}")


DOC_TEXT = (
    "UnityWorks Product Price List 2026.\n\n"
    "The Starter plan costs 99 euros per month and includes 5 users.\n\n"
    "The Professional plan costs 299 euros per month and includes 25 users.\n\n"
    "The Enterprise plan costs 899 euros per month and includes unlimited users "
    "plus dedicated support.\n"
)

with httpx.Client(timeout=600) as c:
    r = c.post(f"{BASE}/auth/register", json={
        "email": EMAIL, "password": PASSWORD, "full_name": "P5 E2E",
        "role": "developer", "org_id": "default",
    })
    if r.status_code not in (200, 201):
        sys.exit(f"REGISTER_FAILED {r.status_code} {r.text[:200]}")
    r = c.post(f"{BASE}/auth/login", json={"email": EMAIL, "password": PASSWORD})
    assert r.status_code == 200, r.text
    H = {"Authorization": f"Bearer {r.json()['access_token']}"}

    # ── Seed knowledge ───────────────────────────────────────────────────────
    r = c.post(f"{BASE}/documents/upload", headers=H, files={
        "file": ("prices.txt", DOC_TEXT.encode(), "text/plain")
    })
    check("upload -> 201", r.status_code == 201, r.text[:200])
    doc_id = r.json()["document"]["id"]
    deadline = time.time() + 90
    sem_ready = False
    while time.time() < deadline:
        rs = c.get(f"{BASE}/documents/{doc_id}/semantic", headers=H)
        if rs.status_code == 200 and rs.json().get("status") == "indexed":
            sem_ready = True
            break
        time.sleep(1.5)
    check("document semantically indexed", sem_ready)

    # ── Supported formats ────────────────────────────────────────────────────
    r = c.get(f"{BASE}/generations/formats", headers=H)
    check("formats -> 200", r.status_code == 200, r.text[:200])
    formats = r.json()["formats"]
    check("all 7 formats supported",
          set(formats) == {"csv", "excel", "html", "json", "markdown", "pdf", "word"},
          str(formats))

    # ── Excel generation: the flagship path (grounded, tabular) ──────────────
    r = c.post(f"{BASE}/generations", headers=H, json={
        "format": "excel",
        "prompt": "Create a pricing overview spreadsheet with a table of all "
                  "UnityWorks plans, their monthly price in euros, and user limits.",
    })
    check("generate excel -> 201", r.status_code == 201, r.text[:300])
    art = r.json()
    check("excel status == ready", art.get("status") == "ready",
          f"status={art.get('status')} error={art.get('error')}")
    check("excel grounded in knowledge", art.get("grounded") is True)
    check("excel builder recorded", art.get("builder_name") == "excel")
    check("checksum is sha256 hex", len(art.get("checksum", "")) == 64)
    check("size_bytes > 0", art.get("size_bytes", 0) > 0)
    check("source knowledge ids recorded",
          art.get("source_knowledge_ids_json") is not None
          and len(json.loads(art["source_knowledge_ids_json"])) >= 1)
    check("llm model recorded", art.get("llm_model") != "")
    check("metrics: planning/build/store/total ms captured",
          all(isinstance(art.get(k), int) for k in
              ("planning_ms", "transform_ms", "build_ms", "store_ms", "total_ms")),
          str({k: art.get(k) for k in ("planning_ms", "build_ms", "store_ms")}))
    check("metrics: real llm tokens", (art.get("prompt_tokens") or 0) > 0)
    check("correlation id present", len(art.get("correlation_id", "")) == 36)
    excel_id = art["id"]

    # ── Download: signed URL (S3 backend) or proxy fallback ─────────────────
    r = c.get(f"{BASE}/generations/{excel_id}/download", headers=H)
    check("download -> 200", r.status_code == 200, r.text[:200])
    ct = r.headers.get("content-type", "")
    excel_bytes = None
    if "application/json" in ct:
        dl = r.json()
        check("download mode == signed_url", dl.get("mode") == "signed_url", str(dl)[:200])
        check("signed url present with ttl", bool(dl.get("url")) and dl.get("expires_in", 0) > 0)
        rf = httpx.get(dl["url"], timeout=60)
        check("signed url fetch -> 200", rf.status_code == 200, str(rf.status_code))
        excel_bytes = rf.content
    else:
        check("download mode == proxy bytes", "spreadsheetml" in ct, ct)
        excel_bytes = r.content
    check("downloaded bytes match registry checksum",
          excel_bytes is not None
          and hashlib.sha256(excel_bytes).hexdigest() == art["checksum"])
    # Validate it is a real xlsx with a real table inside
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(excel_bytes))
        joined = " ".join(
            str(cell) for ws in wb.worksheets for row in ws.iter_rows(values_only=True)
            for cell in row if cell
        )
        check("xlsx opens and contains plan pricing", "99" in joined or "299" in joined,
              joined[:200])
    except Exception as e:
        check("xlsx opens and contains plan pricing", False, str(e))

    # ── Lifecycle events ─────────────────────────────────────────────────────
    r = c.get(f"{BASE}/generations/{excel_id}/events", headers=H)
    check("events -> 200", r.status_code == 200)
    evs = [e["event_type"] for e in r.json()]
    for expected in ("generation_requested", "plan_completed", "transform_completed",
                     "build_completed", "artifact_stored", "artifact_ready",
                     "artifact_downloaded"):
        check(f"event {expected} recorded", expected in evs, str(evs))

    # ── Remaining formats (one LLM plan each, real render + store) ───────────
    for fmt in ("pdf", "word", "csv", "json", "markdown", "html"):
        r = c.post(f"{BASE}/generations", headers=H, json={
            "format": fmt,
            "prompt": "Create a short pricing summary of the UnityWorks plans.",
            "document_id": doc_id,
        })
        ok = r.status_code == 201 and r.json().get("status") == "ready"
        check(f"generate {fmt} -> ready", ok,
              f"{r.status_code} {r.json().get('error') if r.status_code == 201 else r.text[:200]}")

    # ── Registry listing ─────────────────────────────────────────────────────
    r = c.get(f"{BASE}/generations", headers=H)
    check("registry lists all 7 artifacts", r.status_code == 200 and len(r.json()) == 7,
          str(len(r.json()) if r.status_code == 200 else r.status_code))

    # ── Builder capability registry (Objective 19) ───────────────────────────
    r = c.get(f"{BASE}/platform/capabilities", headers=H)
    caps = r.json()["items"]
    gens = [x["name"] for x in caps if x["category"] == "generator"]
    check("all 7 builders in capability registry",
          set(gens) >= {"csv", "excel", "html", "json", "markdown", "pdf", "word"},
          str(gens))

    # ── Failure semantics ────────────────────────────────────────────────────
    r = c.post(f"{BASE}/generations", headers=H, json={
        "format": "powerpoint", "prompt": "anything",
    })
    check("unknown format -> 422", r.status_code == 422, str(r.status_code))
    r = c.get(f"{BASE}/generations/{uuid.uuid4()}", headers=H)
    check("unknown artifact -> 404", r.status_code == 404)
    r = c.get(f"{BASE}/generations/{uuid.uuid4()}/download", headers=H)
    check("unknown artifact download -> 404", r.status_code == 404)

print(f"\n{PASS} passed, {FAIL} failed")
sys.exit(1 if FAIL else 0)
