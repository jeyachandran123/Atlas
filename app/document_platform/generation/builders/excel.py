"""Excel builder (openpyxl). Sections with tables become worksheets; a
summary sheet carries title/metadata and non-tabular sections."""
from __future__ import annotations

import io
import re

from app.document_platform.generation.builders.base import (
    FIXED_BUILD_TIME,
    AbstractFileBuilder,
    BuildError,
)
from app.document_platform.generation.content_model import ContentModel

_SHEET_BAD = re.compile(r"[\\/*?:\[\]]")


class ExcelBuilder(AbstractFileBuilder):
    format_name = "excel"
    extension = "xlsx"
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    features = frozenset({"tables", "multi_sheet", "styling"})

    def build(self, model: ContentModel) -> bytes:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            wb.properties.created = FIXED_BUILD_TIME.replace(tzinfo=None)
            wb.properties.modified = FIXED_BUILD_TIME.replace(tzinfo=None)
            wb.properties.creator = "UnityWorks Generation Platform"

            summary = wb.active
            summary.title = "Summary"
            summary["A1"] = model.title
            summary["A1"].font = Font(bold=True, size=14)
            row = 2
            if model.subtitle:
                summary.cell(row=row, column=1, value=model.subtitle)
                row += 1
            for key, value in model.metadata.items():
                summary.cell(row=row, column=1, value=key).font = Font(bold=True)
                summary.cell(row=row, column=2, value=value)
                row += 1
            row += 1
            for section in model.sections:
                if section.table is not None:
                    continue
                summary.cell(row=row, column=1, value=section.heading).font = Font(bold=True)
                row += 1
                for para in section.paragraphs:
                    summary.cell(row=row, column=1, value=para)
                    row += 1
                for bullet in section.bullets:
                    summary.cell(row=row, column=1, value=f"• {bullet}")
                    row += 1
                row += 1

            used = {"Summary"}
            for i, section in enumerate(model.sections):
                table = section.table
                if table is None:
                    continue
                name = _SHEET_BAD.sub("", table.name or section.heading or f"Sheet{i+1}")[:31] \
                    or f"Sheet{i+1}"
                base, n = name, 2
                while name in used:
                    name = f"{base[:28]}_{n}"
                    n += 1
                used.add(name)
                ws = wb.create_sheet(title=name)
                for col, header in enumerate(table.headers, start=1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                for r, data_row in enumerate(table.rows, start=2):
                    for col, value in enumerate(data_row, start=1):
                        ws.cell(row=r, column=col, value=value)
                for col in range(1, len(table.headers) + 1):
                    ws.column_dimensions[get_column_letter(col)].width = 22

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()
        except BuildError:
            raise
        except Exception as e:
            raise BuildError(f"Excel build failed: {e}") from e
